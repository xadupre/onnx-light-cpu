# Copyright (c) ONNX Project Contributors
#
# SPDX-License-Identifier: Apache-2.0

"""Tests for ``python -m onnx_light_cpu benchmark``."""

import io
import re
import sys
import tempfile
from contextlib import redirect_stderr
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from onnx_light.ext_test_case import ExtTestCase
from openpyxl import load_workbook

from onnx_light_cpu.__main__ import _build_parser, main
from onnx_light_cpu import _benchmark
from onnx_light_cpu._benchmark import (
    normalize_dtypes,
    write_benchmark_markdown,
    write_benchmark_workbook,
)


class TestBenchmarkCli(ExtTestCase):
    def test_parser_accepts_test_and_dtype_lists(self):
        args = _build_parser().parse_args(
            [
                "benchmark",
                "--test",
                "^test_cpu_abs_",
                "^test_cpu_gemm_",
                "--dtype",
                "float32,float64",
                "int64",
                "--onnxruntime",
                "--markdown",
                "results.md",
            ]
        )
        self.assertEqual(args.tests, ["^test_cpu_abs_", "^test_cpu_gemm_"])
        self.assertEqual(args.dtypes, ["float32,float64", "int64"])
        self.assertTrue(args.onnxruntime)
        self.assertEqual(args.markdown, "results.md")
        self.assertEqual(
            normalize_dtypes(args.dtypes),
            ("float32", "float64", "int64"),
        )

    def test_rejects_unknown_dtype(self):
        with self.assertRaisesRegex(ValueError, "unknown dtype"):
            normalize_dtypes(["complex128"])

    def test_rejects_invalid_test_regular_expression(self):
        with self.assertRaisesRegex(re.PatternError, "unterminated"):
            _benchmark.run_backend_benchmark(
                tests=["("],
                dtypes=["all"],
                repeat=1,
                warmup=0,
                max_repeat_time=1.0,
                threads=1,
                with_onnxruntime=False,
            )

    def test_selects_backend_tests_and_dtypes(self):
        cases = [
            SimpleNamespace(name="test_cpu_abs_float32_benchmark", unload=mock.Mock()),
            SimpleNamespace(name="test_cpu_abs_float64_benchmark", unload=mock.Mock()),
            SimpleNamespace(name="test_cpu_gemm_float32_benchmark", unload=mock.Mock()),
        ]
        measured = (
            [{"case": cases[0].name, "duration_s": 1.0}],
            {"case": cases[0].name, "mean_s": 1.0},
        )
        with (
            mock.patch("onnx_light_cpu._benchmark.register_backend_test_cases"),
            mock.patch("onnx_light_cpu._benchmark.register_kernels"),
            mock.patch(
                "onnx_light.onnx.backend.collect_test_cases_by_name",
                return_value=cases,
            ),
            mock.patch(
                "onnx_light_cpu._benchmark._measure_case", return_value=measured
            ) as measure,
        ):
            progress = io.StringIO()
            with redirect_stderr(progress):
                raw, aggregated = _benchmark.run_backend_benchmark(
                    tests=[r"^test_cpu_(abs|log)_"],
                    dtypes=["float32"],
                    repeat=1,
                    warmup=0,
                    max_repeat_time=1.0,
                    threads=1,
                )
        self.assertEqual(raw, measured[0])
        self.assertEqual(aggregated, [measured[1]])
        self.assertIn("[1/1] test_cpu_abs_float32_benchmark", progress.getvalue())
        measure.assert_called_once_with(cases[0], 1, 0, 1.0, 1, False)

    def test_explicit_threads_use_onnxruntime_affinity_default(self):
        case = SimpleNamespace(
            name="test_cpu_abs_float32_benchmark",
            model=SimpleNamespace(
                graph=SimpleNamespace(
                    node=[SimpleNamespace(op_type="Abs")],
                    input=[],
                )
            ),
            data_sets=[],
        )
        with (
            mock.patch("onnx_light.onnx.reference.ReferenceEvaluator") as evaluator,
            mock.patch("onnx_light_cpu._benchmark.clear_used_kernel_names"),
            mock.patch("onnx_light_cpu._benchmark.platform.processor", return_value="test CPU"),
            mock.patch(
                "onnx_light_cpu._benchmark.used_kernel_names",
                return_value=("onnx_light_cpu::Abs",),
            ),
        ):
            raw, _ = _benchmark._measure_case(case, 1, 0, 1.0, 3)
        evaluator.assert_called_once_with(
            case.model,
            cpu_execution={"num_threads": 3, "affinity_policy": "none"},
        )
        self.assertEqual(raw[0]["run"], 1)
        self.assertEqual(raw[0]["runtime"], "onnx-light-cpu")
        self.assertEqual(raw[0]["processor"], "test CPU")
        self.assertIn("duration_s", raw[0])

    def test_onnxruntime_unsupported_case_is_reported(self):
        model = SimpleNamespace(
            graph=SimpleNamespace(
                node=[SimpleNamespace(op_type="Abs")],
                input=[],
            ),
            SerializeToString=mock.Mock(return_value=b"model"),
        )
        case = SimpleNamespace(
            name="test_cpu_abs_float32_benchmark",
            model=model,
            data_sets=[],
        )
        onnxruntime = SimpleNamespace(
            SessionOptions=lambda: SimpleNamespace(),
            ExecutionMode=SimpleNamespace(ORT_SEQUENTIAL=0),
            InferenceSession=mock.Mock(side_effect=RuntimeError("unsupported model")),
        )
        with (
            mock.patch.dict(sys.modules, {"onnxruntime": onnxruntime}),
            mock.patch("onnx_light.onnx.reference.ReferenceEvaluator"),
            mock.patch("onnx_light_cpu._benchmark.clear_used_kernel_names"),
            mock.patch(
                "onnx_light_cpu._benchmark.used_kernel_names",
                return_value=("onnx_light_cpu::Abs",),
            ),
        ):
            raw, aggregated = _benchmark._measure_case(case, 1, 0, 1.0, 1, True)
        self.assertEqual(aggregated["onnxruntime_error"], "unsupported model")
        self.assertIsNone(aggregated["onnxruntime_median_s"])
        self.assertIsNone(aggregated["speedup"])
        self.assertEqual([row["runtime"] for row in raw], ["onnx-light-cpu"])

    def test_raw_rows_identify_run_and_runtime(self):
        model = SimpleNamespace(
            graph=SimpleNamespace(
                node=[SimpleNamespace(op_type="Abs")],
                input=[],
            ),
            SerializeToString=mock.Mock(return_value=b"model"),
        )
        case = SimpleNamespace(
            name="test_cpu_abs_float32_benchmark",
            model=model,
            data_sets=[],
        )
        ort_session = SimpleNamespace(run=mock.Mock())
        onnxruntime = SimpleNamespace(
            SessionOptions=lambda: SimpleNamespace(),
            ExecutionMode=SimpleNamespace(ORT_SEQUENTIAL=0),
            InferenceSession=mock.Mock(return_value=ort_session),
        )
        with (
            mock.patch.dict(sys.modules, {"onnxruntime": onnxruntime}),
            mock.patch("onnx_light.onnx.reference.ReferenceEvaluator"),
            mock.patch("onnx_light_cpu._benchmark.clear_used_kernel_names"),
            mock.patch(
                "onnx_light_cpu._benchmark.used_kernel_names",
                return_value=("onnx_light_cpu::Abs",),
            ),
        ):
            raw, aggregated = _benchmark._measure_case(case, 2, 0, 1.0, 1, True)

        self.assertEqual(
            [(row["run"], row["runtime"]) for row in raw],
            [
                (1, "onnx-light-cpu"),
                (2, "onnx-light-cpu"),
                (1, "onnxruntime"),
                (2, "onnxruntime"),
            ],
        )
        self.assertTrue(all(row["duration_s"] >= 0.0 for row in raw))
        self.assertEqual(aggregated["samples"], 2)
        self.assertEqual(aggregated["onnxruntime_samples"], 2)

    def test_writes_raw_and_aggregated_sheets(self):
        raw = [
            {
                "case": "test_cpu_abs_float32_benchmark",
                "operator": "Abs",
                "dtype": "float32",
                "repeat": 2,
                "warmup": 1,
                "threads": 3,
                "processor": "test CPU",
                "input_shapes": '[{"x": [2, 3]}]',
                "max_repeat_time": 1.0,
                "run": 1,
                "runtime": "onnx-light-cpu",
                "duration_s": 0.0000125,
            }
        ]
        aggregated = [
            {
                "case": "test_cpu_abs_float32_benchmark",
                "operator": "Abs",
                "dtype": "float32",
                "repeat": 2,
                "warmup": 1,
                "threads": 3,
                "processor": "test CPU",
                "input_shapes": '[{"x": [2, 3]}]',
                "max_repeat_time": 1.0,
                "samples": 1,
                "mean_s": 0.0000125,
                "stdev_s": 0.0,
                "min_repeat_s": 0.0000125,
                "p10_s": 0.0000125,
                "median_s": 0.0000125,
                "p90_s": 0.0000125,
                "max_repeat_s": 0.0000125,
                "onnxruntime_samples": None,
                "onnxruntime_mean_s": None,
                "onnxruntime_median_s": None,
                "onnxruntime_error": None,
                "speedup": None,
            }
        ]
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "benchmark.xlsx"
            write_benchmark_workbook(output, raw, aggregated)
            workbook = load_workbook(output, read_only=True)
            self.assertEqual(workbook.sheetnames, ["raw", "aggregated"])
            self.assertEqual(workbook["raw"]["A2"].value, raw[0]["case"])
            raw_headers = [cell.value for cell in workbook["raw"][1]]
            self.assertIn("run", raw_headers)
            self.assertIn("runtime", raw_headers)
            self.assertIn("processor", raw_headers)
            self.assertNotIn("duration_us", raw_headers)
            self.assertEqual(workbook["aggregated"]["J2"].value, 1)
            workbook.close()

    def test_writes_aggregated_markdown(self):
        aggregated = [
            {
                column: "test|value" if column == "case" else 1
                for column in _benchmark._AGGREGATED_COLUMNS
            }
        ]
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "benchmark.md"
            write_benchmark_markdown(output, aggregated)
            self.assertEqual(
                output.read_text(encoding="utf-8"),
                "| case | operator | dtype | repeat | warmup | threads | processor | input_shapes"
                " | max_repeat_time | samples | mean_s | stdev_s | min_repeat_s | p10_s"
                " | median_s | p90_s | max_repeat_s | onnxruntime_samples | onnxruntime_mean_s"
                " | onnxruntime_median_s | onnxruntime_error | speedup |\n"
                "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | ---"
                " | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |\n"
                "| test\\|value | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1"
                " | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 |\n",
            )

    def test_main_runs_benchmark_and_writes_output(self):
        rows = ([{"duration_s": 1.0}], [{"case": "one"}])
        with (
            mock.patch(
                "onnx_light_cpu.__main__.run_backend_benchmark",
                return_value=rows,
            ) as run,
            mock.patch("onnx_light_cpu.__main__.write_benchmark_workbook") as write,
            mock.patch("onnx_light_cpu.__main__.write_benchmark_markdown") as markdown,
        ):
            result = main(
                [
                    "benchmark",
                    "--tests",
                    r"^test_cpu_(abs|log)_",
                    "--dtypes",
                    "float32",
                    "--repeat",
                    "2",
                    "--warmup",
                    "0",
                    "--threads",
                    "1",
                    "--output",
                    "results.xlsx",
                    "--markdown",
                    "results.md",
                ]
            )
        self.assertEqual(result, 0)
        run.assert_called_once_with(
            tests=[r"^test_cpu_(abs|log)_"],
            dtypes=["float32"],
            repeat=2,
            warmup=0,
            max_repeat_time=1.0,
            threads=1,
            with_onnxruntime=False,
        )
        write.assert_called_once_with("results.xlsx", *rows)
        markdown.assert_called_once_with("results.md", rows[1])


if __name__ == "__main__":
    import unittest

    unittest.main()
